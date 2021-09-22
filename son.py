# coding=utf-8
import random

import numpy as np
from openpyxl import Workbook, load_workbook
from array import array
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QDialog, QBoxLayout, QListWidget, QLabel
from PyQt5 import QtCore, QtGui
from PyQt5.QtCore import *
import sys
import PyQt5.QtWidgets as qtw
import PyQt5.QtGui as qtg
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QMessageBox

secenek = ["Bilgisayar", "Cep Telefonu", "Buzdolabı"]
wb = load_workbook("GUNCELDATA.xlsx")
ws = wb.sheetnames
ws = wb['Sayfa1']

# buzdolabı
coolerlist = []
valuescooler = [c.value for c in ws['B'][696:913]]
for a in valuescooler:
    if a not in coolerlist:
        coolerlist.append(a)


# combobox olusturma

class Window(QMainWindow):

    def on_combobox_changed(self, value):
        if value == 'Cep Telefonu':
            self.getirPhone()
        elif value == 'Bilgisayar':
            self.getirpc()
        else:
            self.getirCooler()

    # do your code
    def __init__(self):
        super().__init__()

        # setting title
        self.setWindowTitle("Mesleki Uygulama ")

        # setting geometry
        self.setGeometry(100, 100, 1080, 720)
        Window.setMaximumSize(self, 1440, 1280)
        self.UiComponents()
        self.show()

    # method for widgets
    def UiComponents(self):

        self.kategori = QComboBox(self)

        # setting geometry of combo box
        self.kategori.setGeometry(400, 5, 191, 31)
        self.kategori.setStyleSheet("background-color: rgb(255, 235, 188);")

        self.kategori.addItems(secenek)
        self.kategori.currentTextChanged.connect(self.on_combobox_changed)

        self.button1 = QPushButton("Onayla", self)
        self.button1.setGeometry(430, 50, 93, 28)
        self.button1.setStyleSheet("background-color: rgb(85, 170, 255);")
        self.button1.clicked.connect(self.getirpc)
        # URUNLİST
        self.urunlist = QListWidget(self)
        self.urunlist.setGeometry(30, 160, 341, 541)
        self.urunlist.currentTextChanged.connect(self.itemChanged)

        # urunlıstesı baslık
        self.label1 = QLabel("          ÜrÜn Listesi", self)
        self.label1.setGeometry(130, 90, 161, 41)
        self.label1.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label1.setFrameShape(QFrame.Box)
        # onaylabutonu
        self.button2 = QPushButton("Temizle", self)
        self.button2.setGeometry(420, 340, 93, 28)
        self.button2.setStyleSheet("background-color: rgb(85, 170, 255);")
        self.button2.clicked.connect(self.temizle)

        self.label2 = QLabel("         Yorumlar", self)
        self.label2.setGeometry(700, 100, 131, 41)
        self.yorum = QListWidget(self)
        self.yorum.setGeometry(560, 170, 421, 251)
        self.label2.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label2.setFrameShape(QFrame.Box)

        # İstatistik
        self.label3 = QLabel("          İstatistik", self)
        self.label3.setGeometry(710, 440, 131, 41)
        self.label3.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label3.setFrameShape(QFrame.Box)
        self.yuzde = QListWidget(self)
        self.yuzde.setGeometry(560, 500, 431, 201)

    def temizle(self):
        self.urunlist.clear()
        self.yorum.clear()
        self.yuzde.clear()
    def yorumEkle(self,value):
        self.yorum.addItem(value)

    def yuzdeEkle(self,value):
        self.yuzde.addItem(str(value))

    def itemChanged(self, value):
        olumlu  = str(random.randint(7, 20)) #6
        olumsuz  = str(random.randint(1, 7))#3
        yorum = "Olumlu Yorum sayisi : "+ str(olumlu) + "\n Olumsuz yorum sayisi : " +str(olumsuz)
        yuzde = float(int(olumlu) - int(olumsuz) / int(int(olumlu+olumsuz)*100))

        print(yuzde)
        self.yorumEkle(yorum)
        self.yuzdeEkle(yuzde)
        # for c in ws['A'][1:9]: #hepsini al
            # if c.value == value:
            #     print(c)
            # print(c)

            # msg = QMessageBox()
            # msg.setIcon(QMessageBox.Critical)
            # msg.setText("Error")
            # msg.setInformativeText(c.value)
            # msg.setWindowTitle("oldu")
            # msg.exec_()
            # if c['B'] == value:
            #     return c['J']
    #
    # def itemChanged(self,value):
    #     print(value)

    def getirpc(self):
        self.urunlist.clear()
        pclist = []
        valuespc = [c.value for c in ws['B'][1:229]]
        for i in valuespc[:240]:
            if i not in pclist:
                pclist.append(i)
        for a in pclist:
            self.urunlist.addItem(a)

    def getirPhone(self):
        self.urunlist.clear()
        phonelist = []
        valuespc = [c.value for c in ws['B'][237:669]]
        for i in valuespc[:240]:
            if i not in phonelist:
                phonelist.append(i)
        for a in phonelist:
            self.urunlist.addItem(a)

    def getirCooler(self):
        self.urunlist.clear()
        coolerlist = []
        valuespc = [c.value for c in ws['B'][705:923]]
        for i in valuespc[:240]:
            if i not in coolerlist:
                coolerlist.append(i)
        for a in coolerlist:
            self.urunlist.addItem(a)


app = QApplication(sys.argv)
screen = Window()
screen.show()
sys.exit(app.exec_())

# create pyqt5 app
App = QApplication(sys.argv)

# create the instance of our Window
window = Window()

# start the app
sys.exit(App.exec())
