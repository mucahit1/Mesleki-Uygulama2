# coding=utf-8
import random

import numpy as np
from openpyxl import Workbook, load_workbook
from array import array
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QDialog, QBoxLayout, QListWidget, QLabel
from PyQt5 import QtCore, QtGui
from PyQt5.QtCore import *
import sys
import PyQt5.QtWidgets as qtw
import PyQt5.QtGui as qtg
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import QMessageBox

secenek = ["Bilgisayar", "Cep Telefonu", "Buzdolabı"]
wb = load_workbook("GUNCELDATA.xlsx")
ws = wb.sheetnames
ws = wb['Sayfa1']


# combobox olusturma

class Window(QMainWindow):

    def on_combobox_changed(self, value):
        if value == 'Cep Telefonu':
            self.getirPhone()
        elif value == 'Bilgisayar':
            self.getirpc()
        elif value=="Buzdolabı":
            self.getirCooler()


    # do your code
    def __init__(self):
        super().__init__()

        # setting title
        self.setWindowTitle("Mesleki Uygulama ")
        self.setWindowIcon(QtGui.QIcon('logo1.png'))
        self.setStyleSheet("rgb(228, 236, 255)")
        # setting geometry
        self.setGeometry(100, 100, 1080, 720)
        Window.setMaximumSize(self, 1280, 920)
        self.UiComponents()
        self.show()

    # method for widgets
    def UiComponents(self):

        self.kategori = QComboBox(self)

        # setting geometry of combo box
        self.kategori.setGeometry(400, 5, 191, 31)
        self.kategori.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.kategori.addItems(secenek)
        self.kategori.currentTextChanged.connect(self.on_combobox_changed)

        self.button1 = QPushButton("Onayla", self)
        self.button1.setGeometry(430, 50, 93, 28)
        self.button1.setStyleSheet("background-color: rgb(85, 170, 255);")


        # URUNLİST
        self.urunlist = QListWidget(self)
        self.urunlist.setGeometry(30, 160, 341, 541)
        self.urunlist.currentTextChanged.connect(self.listChanged)

        # urunlıstesı baslık
        self.label1 = QLabel("          ÜrÜn Listesi", self)
        self.label1.setGeometry(130, 90, 161, 41)
        self.label1.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label1.setFrameShape(QFrame.Box)
        # onaylabutonu
        self.button2 = QPushButton("Temizle", self)
        self.button2.setGeometry(420, 340, 93, 28)
        self.button2.setStyleSheet("background-color: rgb(85, 170, 255);")
        self.button2.clicked.connect(self.temizle)
#yorumlar
        self.label2 = QLabel("         Yorumlar", self)
        self.label2.setGeometry(700, 100, 131, 41)
        self.yorum = QListWidget(self)
        self.yorum.setGeometry(560, 170, 421, 251)
        self.label2.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label2.setFrameShape(QFrame.Box)


        # İstatistik
        self.label3 = QLabel("          İstatistik", self)
        self.label3.setGeometry(710, 440, 131, 41)
        self.label3.setStyleSheet("background-color: rgb(255, 235, 188);")
        self.label3.setFrameShape(QFrame.Box)
        self.yuzde = QListWidget(self)
        self.yuzde.setGeometry(560, 500, 431, 201)

    def temizle(self):
        self.urunlist.clear()
        self.yorum.clear()
        self.yuzde.clear()
    def yorumEkle(self,value):
        self.yorum.clear()
        self.yorum.addItem(value)

    def yuzdeEkle(self,value):
        self.yuzde.clear()
        self.yuzde.addItem(str(value))



    def getirpc(self):
        self.urunlist.clear()
        pclist = []
        valuespc = [c.value for c in ws['B'][1:229]]
        for i in valuespc[:229]:
            if i not in pclist:
                pclist.append(i)
        for a in pclist:
            self.urunlist.addItem(a)

    def getirPhone(self):
        self.urunlist.clear()
        phonelist = []
        valuespc = [c.value for c in ws['B'][237:669]]
        for i in valuespc[:240]:
            if i not in phonelist:
                phonelist.append(i)
        for a in phonelist:
            self.urunlist.addItem(a)


    def getirCooler(self):
        self.urunlist.clear()
        coolerlist = []
        valuespc = [c.value for c in ws['B'][705:923]]
        for i in valuespc[:240]:
            if i not in coolerlist:
                coolerlist.append(i)
        for a in coolerlist:
            self.urunlist.addItem(a)

    def listChanged(self, value):

        if value =='Lenovo IdeaPad Gaming':
            self.yorumEkle("Urun puanı: 4.8"+ "\n" "Olumlu Yorum sayısı:  20" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value =='Lenovo IdeaPad AMD Ryzen':
            self.yorumEkle("Urun puanı: 4.7"+ "\n" "Olumlu Yorum sayısı:  10" "\n" + "Olumsuz Yorum Sayısı:   7")
            self.yuzdeEkle("Ürün Başarım Oranı: %58" "\n")

        elif value=="HP 15SFQ2045NT ":
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  20" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value==" Dell Vostro 3501 ":
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %79""\n")

        elif value=="Acer Swift SF31442 ":
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  16" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %80""\n")

        elif value=='Asus TUF FX506LIHN085':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  16" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %20""\n")

        elif value=='Lenovo V15IIL':
            self.yorumEkle("Urun puanı: 4.75" + "\n" "Olumlu Yorum sayısı:  18" "\n" + "Olumsuz Yorum Sayısı:   2")
            self.yuzdeEkle("Ürün Başarım Oranı: %90""\n")

        elif value=='Asus D509DAEJ887':
            self.yorumEkle("Urun puanı: 4.8" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %79""\n")

        elif value=='Lenovo V15':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  14" "\n" + "Olumsuz Yorum Sayısı:   6")
            self.yuzdeEkle("Ürün Başarım Oranı: %70""\n")

        elif value=='Lenovo IdeaPad 3  3700U':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  17" "\n" + "Olumsuz Yorum Sayısı:   3")
            self.yuzdeEkle("Ürün Başarım Oranı: %85""\n")

        elif value=='I-Life Zed Air':
            self.yorumEkle("Urun puanı: 4.1" + "\n" "Olumlu Yorum sayısı:  8" "\n" + "Olumsuz Yorum Sayısı:   12")
            self.yuzdeEkle("Ürün Başarım Oranı: %40""\n")

        elif value == 'Lenovo IdeaPad 5  5700U':
            self.yorumEkle("Urun puanı: 4" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   5")
            self.yuzdeEkle("Ürün Başarım Oranı: %75""\n")
##PHONE
        elif value=='Samsung Galaxy M31 2020 128 GB':
            self.yorumEkle("Urun puanı:  4.8" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %93.7""\n")


        elif value=='iPhone 11 64 GB':
            self.yorumEkle("Urun puanı: 4.8" + "\n" "Olumlu Yorum sayısı:  19" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %95""\n")

        elif value =='Redmi Note9 Pro128GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  19" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100")

        elif value=='Xiaomi Redmi Note 8 64 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  20" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")


        elif value=='Samsung Galaxy M51 128 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  20" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='Xiaomi Redmi Note 8 Pro 64 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  18" "\n" + "Olumsuz Yorum Sayısı:   2")
            self.yuzdeEkle("Ürün Başarım Oranı: %90""\n")


        elif value=='iPhone 11 128 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  10" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='Xiaomi Redmi Note 9 128 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  11" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %91.6""\n")

        elif value=='Xiaomi Redmi Note 8 Pro 128 GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  21" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")


        elif value=='Samsung Galaxy M21 64GB':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  16" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %94.1""\n")

        elif value=='Samsung Galaxy M20 32 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   3")
            self.yuzdeEkle("Ürün Başarım Oranı: %83.3""\n")

        elif value=='Xiaomi Mi Note 10 Lite 128 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  4" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %80""\n")


        elif value=='Samsung Galaxy A51 2020 128 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  5" "\n" + "Olumsuz Yorum Sayısı:   2")
            self.yuzdeEkle("Ürün Başarım Oranı: %71.4""\n")

        elif value=='Oppo Reno2 Z 128 GB ':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  3" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %75""\n")


        elif value=='Xiaomi Redmi Note 9 Pro 64 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  6" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %85""\n")


        elif value=='Samsung Galaxy S10 Plus 128 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  4" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='iPhone SE 64 GB':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  17" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %94.4""\n")

        elif value=='Samsung Galaxy M31S 128 GB':
            self.yorumEkle("Urun puanı: 4.4" + "\n" "Olumlu Yorum sayısı:  18" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='Samsung Galaxy A71 2020 128 GB':
            self.yorumEkle("Urun puanı: 4.4" + "\n" "Olumlu Yorum sayısı:  17" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %94.4""\n")

        elif value=='Samsung Galaxy Note 10 Lite 128 GB':
            self.yorumEkle("Urun puanı: 4.4" + "\n" "Olumlu Yorum sayısı:  16" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='Vestel SC470 A+ 470 lt Statik Buzdolabı':
            self.yorumEkle("Urun puanı: 4.7" + "\n" "Olumlu Yorum sayısı:  17" "\n" + "Olumsuz Yorum Sayısı:   0")
            self.yuzdeEkle("Ürün Başarım Oranı: %100""\n")

        elif value=='Vestel 20263685 NF48001 Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  6" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %85.7""\n")

        elif value == 'Vestfrost VF 1268 300 lt Statik Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   5")
            self.yuzdeEkle("Ürün Başarım Oranı: %75""\n")


        elif value=='Samsung RT46K6000WW/TR 468 lt No-Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  10" "\n" + "Olumsuz Yorum Sayısı:   6")
            self.yuzdeEkle("Ürün Başarım Oranı: %62.5""\n")

        elif value=='Altus AL-306 E A+ 140 lt Statik Büro Tipi Mini Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  25" "\n" + "Olumsuz Yorum Sayısı:   2")
            self.yuzdeEkle("Ürün Başarım Oranı: %92.5""\n")

        elif value=='Samsung RT46K6000S8/TR 468 lt No-Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  18" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %94.7""\n")

        elif value=='Altus AL-370 N A+ 465 lt No-Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  19" "\n" + "Olumsuz Yorum Sayısı:   9")
            self.yuzdeEkle("Ürün Başarım Oranı: %67.8""\n")

        elif value=='Bosch KDN55NWF0N 485 lt No Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.6" + "\n" "Olumlu Yorum sayısı:  19" "\n" + "Olumsuz Yorum Sayısı:   1")
            self.yuzdeEkle("Ürün Başarım Oranı: %95""\n")

        elif value=='LG GN-H702HLHU 546 lt No-Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  14" "\n" + "Olumsuz Yorum Sayısı:   6")
            self.yuzdeEkle("Ürün Başarım Oranı: %70""\n")


        elif value=='Dijitsu Db 100 Büro Tipi Mini Buzdolabı':
            self.yorumEkle("Urun puanı: 4.5" + "\n" "Olumlu Yorum sayısı:  " "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %33.3""\n")

        elif value=='Samsung RT38K50AJWW/TR 401 lt No Frost Buzdolabı':
            self.yorumEkle("Urun puanı: 4.4" + "\n" "Olumlu Yorum sayısı:  15" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %79""\n")

        elif value=='Altus Al 306 B 114 Lt Buzdolabı':
            self.yorumEkle("Urun puanı: 4.4" + "\n" "Olumlu Yorum sayısı:  16" "\n" + "Olumsuz Yorum Sayısı:   4")
            self.yuzdeEkle("Ürün Başarım Oranı: %80""\n")





        elif value=='':
            self.yorumEkle("Urun puanı: 4.8" + "\n" "Olumlu Yorum sayısı:  4" "\n" + "Olumsuz Yorum Sayısı:   15")
            self.yuzdeEkle("Ürün Başarım Oranı: %79""\n")





app = QApplication(sys.argv)
screen = Window()
screen.show()
sys.exit(app.exec_())

# create pyqt5 app
App = QApplication(sys.argv)

# create the instance of our Window
window = Window()

# start the app
sys.exit(App.exec())
